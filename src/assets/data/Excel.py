import cgi
import cgitb

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series


wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

values = Reference(ws1, min_col=1, min_row=1, max_col=1, max_row=10)
chart = BarChart()
chart.add_data(values)
ws1.add_chart(chart, "E15")

ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))

wb.save(filename=dest_filename)

print dest_filename + " created"
